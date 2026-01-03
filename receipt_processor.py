import pandas as pd
import shutil
import logging
import xlwt
from pathlib import Path
from datetime import datetime
from typing import List, Generator, Dict, Optional, Tuple, Any
import time
import os
import logger_config  # 導入統一的日誌配置

# 使用統一的日誌配置
logger = logging.getLogger(__name__)

# --- 設定管理器 (單一表格模式) ---
class ConfigManager:
    def __init__(self, base_dir: Path):
        self.settings_dir = base_dir / "settings"
        self.settings_file = self.settings_dir / "supplier_config.xlsx"
        self.settings_dir.mkdir(parents=True, exist_ok=True)
        
        # 預設範例 (依照你的要求：供應商名稱只是顯示用)
        self.DEFAULT_DATA = [
            {
                "供應商名稱": "可美塑料制品有限公司", # 這是給人看的 ID
                "貨品條碼": "条形码",             # 必須完全匹配收據裡的欄位名
                "入貨價": "单价",
                "入貨量": "数量",
                "貨品名稱": "货品名称及规格型号",
                "成本乘以1.2": "是"
            }
        ]

    def load_config(self) -> pd.DataFrame:
        """讀取 Excel 設定"""
        if not self.settings_file.exists():
            self._create_default_config()
        
        try:
            df = pd.read_excel(self.settings_file)
            # 確保欄位都是字串，並去除前後空白
            df = df.astype(str).apply(lambda x: x.str.strip())
            logger.info("✅ 供應商設定檔讀取成功")
            return df
        except Exception as e:
            logger.error(f"❌ 讀取設定檔失敗: {e}")
            return pd.DataFrame(self.DEFAULT_DATA)

    def _create_default_config(self):
        """產生給員工填寫的 Excel"""
        logger.info(f"⚠️ 建立預設設定檔: {self.settings_file}")
        df = pd.DataFrame(self.DEFAULT_DATA)
        cols = ["供應商名稱", "貨品條碼", "入貨價", "入貨量", "貨品名稱", "成本乘以1.2"]
        df = df[cols]
        df.to_excel(self.settings_file, index=False)


# --- Mapping 管理器 ---
class MappingManager:
    def __init__(self, base_dir: Path):
        """
        管理條碼到貨品編號的映射關係
        
        Args:
            base_dir: 工作目錄路徑
        """
        self.settings_dir = base_dir / "settings"
        self.mapping_file = self.settings_dir / "barcode_mapping.xlsx"
        self.settings_dir.mkdir(parents=True, exist_ok=True)
        self.mapping_df = self.load_mapping()
    
    def load_mapping(self) -> pd.DataFrame:
        """讀取現有的 mapping"""
        if not self.mapping_file.exists():
            # 建立空的 mapping 檔案
            df = pd.DataFrame(columns=['貨品條碼', '貨品名稱', '貨品編號', '供應商名稱', '建立日期'])
            df.to_excel(self.mapping_file, index=False)
            logger.info(f"📝 建立新的 mapping 檔案: {self.mapping_file}")
            return df
        
        try:
            df = pd.read_excel(self.mapping_file)
            # 確保欄位都是字串，並去除前後空白
            df = df.astype(str).apply(lambda x: x.str.strip())
            logger.info(f"✅ 已載入 mapping 記錄: {len(df)} 筆")
            return df
        except Exception as e:
            logger.error(f"❌ 讀取 mapping 檔案失敗: {e}")
            return pd.DataFrame(columns=['貨品條碼', '貨品名稱', '貨品編號', '供應商名稱', '建立日期'])
    
    def find_mapping(self, barcode: str, product_name: str) -> Optional[str]:
        """
        查找是否有對應的 mapping
        
        Args:
            barcode: 貨品條碼
            product_name: 貨品名稱
        
        Returns:
            對應的貨品編號，如果找不到則返回 None
        """
        if self.mapping_df.empty:
            return None
        
        # 清洗條碼格式（與驗證邏輯一致）
        barcode_clean = pd.Series([str(barcode)]).str.strip().str.replace(r'\.0+$', '', regex=True).iloc[0]
        product_name_clean = str(product_name).strip()
        
        # 查找匹配的記錄
        mask = (
            (self.mapping_df['貨品條碼'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True) == barcode_clean) &
            (self.mapping_df['貨品名稱'].astype(str).str.strip() == product_name_clean)
        )
        
        matched = self.mapping_df[mask]
        if not matched.empty:
            product_code = matched.iloc[0]['貨品編號']
            logger.debug(f"   🔍 找到 mapping: {barcode_clean} -> {product_code}")
            return str(product_code).strip()
        
        return None
    
    def add_mapping(self, barcode: str, product_name: str, product_code: str, supplier_name: str = ""):
        """
        新增 mapping 記錄
        
        Args:
            barcode: 貨品條碼
            product_name: 貨品名稱
            product_code: 貨品編號
            supplier_name: 供應商名稱
        """
        # 檢查是否已存在
        barcode_clean = pd.Series([str(barcode)]).str.strip().str.replace(r'\.0+$', '', regex=True).iloc[0]
        product_name_clean = str(product_name).strip()
        
        mask = (
            (self.mapping_df['貨品條碼'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True) == barcode_clean) &
            (self.mapping_df['貨品名稱'].astype(str).str.strip() == product_name_clean)
        )
        
        if mask.any():
            # 更新現有記錄
            self.mapping_df.loc[mask, '貨品編號'] = str(product_code).strip()
            self.mapping_df.loc[mask, '供應商名稱'] = str(supplier_name).strip()
            self.mapping_df.loc[mask, '建立日期'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            logger.info(f"   📝 更新 mapping: {barcode_clean} -> {product_code}")
        else:
            # 新增記錄
            new_row = {
                '貨品條碼': barcode_clean,
                '貨品名稱': product_name_clean,
                '貨品編號': str(product_code).strip(),
                '供應商名稱': str(supplier_name).strip(),
                '建立日期': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.mapping_df = pd.concat([self.mapping_df, pd.DataFrame([new_row])], ignore_index=True)
            logger.info(f"   ➕ 新增 mapping: {barcode_clean} -> {product_code}")
        
        # 儲存到檔案
        self.save_mapping()
    
    def save_mapping(self):
        """儲存 mapping 到檔案"""
        try:
            self.mapping_df.to_excel(self.mapping_file, index=False)
            logger.debug(f"💾 Mapping 已儲存: {len(self.mapping_df)} 筆")
        except Exception as e:
            logger.error(f"❌ 儲存 mapping 失敗: {e}")


# --- 讀取器 ---
class BatchReceiptLoader:
    def __init__(self, base_dir: str = "workspace"):
        self.base_path = Path(base_dir)
        self.pending_path = self.base_path / "pending"
        self.processed_path = self.base_path / "processed"

        # 檢查並建立資料夾
        if not self.pending_path.exists() or not self.processed_path.exists():
            self.pending_path.mkdir(parents=True, exist_ok=True)
            self.processed_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"📂 工作目錄設定完成：")
            logger.info(f"   - 待處理: {self.pending_path}")
            logger.info(f"   - 已歸檔: {self.processed_path}")


    def get_pending_files(self) -> Generator[Path, None, None]:
        """
        掃描 'pending' 資料夾，找出所有 Excel (.xls, .xlsx) 和 CSV 檔案
        """
        # 支援的副檔名
        extensions = ['*.xlsx', '*.xls', '*.csv']
        files_found = False
        
        for ext in extensions:
            # glob 會搜尋所有符合副檔名的檔案
            for file_path in self.pending_path.glob(ext):
                # 忽略 Excel 打開時產生的暫存檔 (檔名以 ~$ 開頭)
                if not file_path.name.startswith('~$'):
                    files_found = True
                    yield file_path
        
        if not files_found:
            logger.warning("⚠️ 'pending' 資料夾內沒有 Excel 或 CSV 檔案")

    def archive_file(self, file_path: Path):
        """歸檔原始文件 (處理檔案被佔用問題)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{file_path.stem}_{timestamp}{file_path.suffix}"
        destination = self.processed_path / new_name

        while True:
            try:
                shutil.move(str(file_path), str(destination))
                logger.info(f"📦 已歸檔: {new_name}")
                break # 成功移動後跳出迴圈
            except PermissionError:
                logger.warning(f"⚠️ 無法移動檔案 (被佔用): {file_path.name}")
                logger.error(f"🛑 錯誤：檔案 '{file_path.name}' 正被 Excel 開啟中！")
                logger.info("👉 請關閉該檔案，然後按 [Enter] 鍵重試...")
                input()  # 等待用戶輸入，但不輸出到終端（通過 logger 已記錄）
                logger.info("🔄 使用者嘗試重試歸檔...")
            except Exception as e:
                logger.error(f"❌ 歸檔失敗 (未知錯誤): {e}")
                break # 其他錯誤直接放棄，避免無窮迴圈

    def smart_load(self, file_path: Path, expected_keywords: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
        try:
            logger.info(f"📖 讀取: {file_path.name}")
            # 1. 先讀前 20 行用來找 Header
            if file_path.suffix == '.csv':
                 df_raw = pd.read_csv(file_path, header=None, nrows=20)
            else:
                df_raw = pd.read_excel(file_path, header=None, nrows=20)
            
            header_idx = self._find_header_row(df_raw, expected_keywords)
            
            if header_idx == -1:
                logger.warning(f"⚠️ {file_path.name}: 找不到標題列，跳過")
                return pd.DataFrame(), pd.DataFrame()

            # 2. 正式讀取數據
            # 修改：加入 dtype=str 強制所有欄位以「純文字」讀取，保留開頭的 0，避免變成數字
            if file_path.suffix == '.csv':
                df_data = pd.read_csv(file_path, header=header_idx, dtype=str)
            else:
                df_data = pd.read_excel(file_path, header=header_idx, dtype=str)
                
            df_data = df_data.dropna(how='all', axis=0).dropna(how='all', axis=1)
            return df_raw, df_data

        except Exception as e:
            logger.error(f"❌ 讀取錯: {e}")
            return pd.DataFrame(), pd.DataFrame()

    def _find_header_row(self, df: pd.DataFrame, keywords: List[str]) -> int:
        max_score = 0
        best_idx = -1
        for idx, row in df.iterrows():
            row_str = " ".join(row.astype(str)).lower()
            score = 0
            for key in keywords:
                # 只要命中其中一個關鍵字
                if key and key.lower() in row_str:
                    score += 1
            
            # 修正：必須命中至少 2 個關鍵字才算找到
            if score > max_score and score >= 2:
                max_score = score
                best_idx = idx
        return best_idx




# --- 清洗器 ---
class ReceiptCleaner:
    def __init__(self, config_df: pd.DataFrame):
        self.config_df = config_df
        
    def identify_supplier_by_columns(self, file_columns: List[str]) -> Tuple[Optional[pd.Series], str]:
        """
        核心邏輯：檢查收據的標題列，是否包含設定檔中某家廠商的所有 4 個關鍵欄位
        """
        if self.config_df.empty:
            return None, "Unknown"

        file_cols_lower = {str(c).strip().lower() for c in file_columns}

        for _, row in self.config_df.iterrows():
            required_cols = [
                str(row['貨品條碼']).strip().lower(),
                str(row['入貨價']).strip().lower(),
                str(row['入貨量']).strip().lower(),
                str(row['貨品名稱']).strip().lower()
            ]
            supplier_name = row['供應商名稱']

            # 檢查是否 4 個欄位都存在
            if set(required_cols).issubset(file_cols_lower):
                logger.info(f"   🎯 欄位完全匹配 -> 識別為: {supplier_name}")
                return row, supplier_name
        
        return None, "New Supplier"

    def process(self, df: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], str]:
        """
        處理收據數據
        
        Returns:
            Tuple[Optional[pd.DataFrame], str]: 
            - DataFrame: 清洗後的數據，如果處理失敗則返回 None
            - str: 識別到的供應商名稱，如果無法識別則返回空字串
        """
        # 1. 識別
        supplier_config, supplier_name = self.identify_supplier_by_columns(df.columns)
        
        # 2. 改名 (只有識別成功才改名，不亂猜)
        if supplier_config is not None:
            df = self._rename_columns_strict(df, supplier_config)
        else:
            logger.warning("   ⚠️ 無法識別供應商 (欄位特徵不符)")
            logger.info(f"      收據欄位: {list(df.columns)}")
            return None, "" # 直接返回 None，不繼續處理

        # 3. 檢查必要欄位
        required_cols = ["貨品條碼", "入貨價", "入貨量", "貨品名稱"]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            logger.error(f"❌ 缺少必要欄位: {missing}")
            return None, supplier_name

        df = df[required_cols].copy()
        
        # 4. 基礎清洗
        # 修改：正則表達式改為 r'\.0+$'，意思是「小數點後跟著一個或多個 0」都要刪掉
        # 這樣可以同時處理 .0, .00, .000 等情況
        df['貨品條碼'] = df['貨品條碼'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True)
        for col in ['入貨價', '入貨量']:
            df[col] = df[col].astype(str).str.replace(r'[^\d.]', '', regex=True)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        df['入貨量'] = df['入貨量'].astype(int)
        
        # 5. 成本加成邏輯
        multiplier_flag = str(supplier_config.get('成本乘以1.2', '')).strip()
        if multiplier_flag in ['是', 'Yes', 'TRUE', 'True', '1']:
            logger.info("   💰 執行成本加成 (x1.2)")
            df['入貨價'] = df['入貨價'] * 1.2
            df['入貨價'] = df['入貨價'].round(2)

        # 6. 補齊欄位
        df['供應商名稱'] = supplier_name  # 填入識別到的供應商名稱
        df['店號'] = 'S1'
        df['入貨日期'] = datetime.now().strftime('%Y%m%d')
        df['收據單號'] = ''
        df['供應商編號'] = '001'
        df['備註'] = ''
        df['狀態'] = ''
        df['貨品編號'] = df['貨品條碼']

        df = df[ (df['貨品條碼'] != '') & (df['貨品條碼'] != 'nan') & (df['入貨量'] > 0) ]
        return df, supplier_name

    def _rename_columns_strict(self, df: pd.DataFrame, config: pd.Series) -> pd.DataFrame:
        """根據設定檔精確改名"""
        target_map = {
            "貨品條碼": config.get("貨品條碼"),
            "入貨價": config.get("入貨價"),
            "入貨量": config.get("入貨量"),
            "貨品名稱": config.get("貨品名稱")
        }

        reverse_map = {}
        for target, source in target_map.items():
            if pd.notna(source):
                reverse_map[str(source).strip().lower()] = target
        
        new_columns = {}
        for col in df.columns:
            if str(col).strip().lower() in reverse_map:
                new_columns[col] = reverse_map[str(col).strip().lower()]
        
        return df.rename(columns=new_columns)


# --- 產品驗證器 ---
class ProductValidator:
    def __init__(self, stock_csv_path: str, mapping_manager: Optional['MappingManager'] = None):
        """
        初始化產品驗證器，讀取 POS 庫存記錄
        
        Args:
            stock_csv_path: DetailGoodsStockToday.csv 的路徑
            mapping_manager: MappingManager 實例，用於檢查已記錄的 mapping
        """
        self.stock_csv_path = Path(stock_csv_path)
        self.productcode_set = set()
        self.barcode_set = set()
        self.mapping_manager = mapping_manager
        self.stock_df = None  # 儲存完整的庫存數據，用於查找共用條碼的選項
        self._load_stock_data()
    
    def _load_stock_data(self):
        """讀取庫存 CSV 並建立查找集合"""
        if not self.stock_csv_path.exists():
            logger.warning(f"⚠️ 庫存檔案不存在: {self.stock_csv_path}")
            return
        
        try:
            # 讀取 CSV，使用字串類型避免格式問題
            df = pd.read_csv(self.stock_csv_path, dtype=str, encoding='utf-8-sig')
            self.stock_df = df  # 儲存完整數據
            
            # 提取 ProductCode 和 Barcode 欄位
            if 'ProductCode' in df.columns:
                # 清洗並轉換為集合：去除空白、處理 .0 結尾、過濾空值
                productcodes = df['ProductCode'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True)
                self.productcode_set = {code for code in productcodes if code and code.lower() != 'nan'}
            
            if 'Barcode' in df.columns:
                # 清洗並轉換為集合：去除空白、處理 .0 結尾、過濾空值
                barcodes = df['Barcode'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True)
                self.barcode_set = {code for code in barcodes if code and code.lower() != 'nan'}
            
            logger.info(f"✅ 已載入庫存記錄: ProductCode {len(self.productcode_set)} 筆, Barcode {len(self.barcode_set)} 筆")
            
        except Exception as e:
            logger.error(f"❌ 讀取庫存檔案失敗: {e}")
    
    def get_barcode_options(self, barcode: str) -> List[Dict[str, str]]:
        """
        獲取共用條碼對應的所有 ProductCode 選項
        
        Args:
            barcode: 條碼
        
        Returns:
            列表，每個元素包含 ProductCode 和 Name
        """
        if self.stock_df is None or 'Barcode' not in self.stock_df.columns:
            return []
        
        barcode_clean = pd.Series([str(barcode)]).str.strip().str.replace(r'\.0+$', '', regex=True).iloc[0]
        
        # 查找所有匹配的記錄
        mask = self.stock_df['Barcode'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True) == barcode_clean
        matched = self.stock_df[mask]
        
        options = []
        for _, row in matched.iterrows():
            product_code = str(row.get('ProductCode', '')).strip()
            name = str(row.get('Name', '')).strip()
            if product_code and product_code.lower() != 'nan':
                options.append({'ProductCode': product_code, 'Name': name})
        
        return options
    
    def validate_products(self, df: pd.DataFrame, supplier_name: str = "") -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        驗證產品是否存在於 POS 系統中
        
        Args:
            df: 清洗後的收據 DataFrame，必須包含「貨品條碼」欄位
            supplier_name: 供應商名稱，用於 mapping 查找
        
        Returns:
            Tuple[pd.DataFrame, pd.DataFrame]: 
            - matched_df: 找到 ProductCode 對應的記錄（正常處理）
            - unmatched_df: 找不到對應的記錄，包含「處理原因」欄位
        """
        if df.empty or '貨品條碼' not in df.columns:
            return pd.DataFrame(), pd.DataFrame()
        
        # 複製 DataFrame 以避免修改原始資料
        df = df.copy()
        
        # 準備比對用的條碼（清洗格式）
        df['_barcode_clean'] = df['貨品條碼'].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True)
        
        # 初始化匹配狀態和原因字典
        matched_mask = pd.Series([False] * len(df), index=df.index)
        reason_dict = {}  # 使用字典儲存每個索引對應的原因
        
        # 統計用
        mapping_count = 0
        matched_count = 0
        barcode_only_count = 0
        unmatched_count = 0
        
        for idx, row in df.iterrows():
            barcode = row['_barcode_clean']
            product_name = str(row.get('貨品名稱', '')).strip()
            
            # 跳過空值
            if not barcode or barcode.lower() == 'nan':
                continue
            
            # 優先檢查 mapping（如果有的話）
            if self.mapping_manager:
                mapped_product_code = self.mapping_manager.find_mapping(barcode, product_name)
                if mapped_product_code:
                    # 找到 mapping，直接使用
                    matched_mask[idx] = True
                    # 更新貨品編號
                    df.loc[idx, '貨品編號'] = mapped_product_code
                    mapping_count += 1
                    continue
            
            # 情況1: 找到 ProductCode（完全匹配）
            if barcode in self.productcode_set:
                matched_mask[idx] = True
                matched_count += 1
            # 情況2: 只找到 Barcode（部分匹配）
            elif barcode in self.barcode_set:
                reason_dict[idx] = '共用條碼，需人手選擇顏色或大小'
                barcode_only_count += 1
            # 情況3: 完全找不到
            else:
                reason_dict[idx] = '可能是條碼錯誤或新貨品'
                unmatched_count += 1
        
        # 分離數據
        matched_df = df[matched_mask].copy()
        unmatched_df = df[~matched_mask].copy()
        
        # 移除臨時欄位
        if '_barcode_clean' in matched_df.columns:
            matched_df = matched_df.drop(columns=['_barcode_clean'])
        if '_barcode_clean' in unmatched_df.columns:
            unmatched_df = unmatched_df.drop(columns=['_barcode_clean'])
        
        # 為 unmatched_df 加入處理原因欄位
        if not unmatched_df.empty:
            # 使用索引對應的原因
            unmatched_df['處理原因'] = unmatched_df.index.map(reason_dict).fillna('')
        
        # 記錄統計
        logger.info(f"   📊 產品驗證結果:")
        if mapping_count > 0:
            logger.info(f"      🔄 使用 Mapping: {mapping_count} 筆")
        logger.info(f"      ✅ 找到 ProductCode: {matched_count} 筆")
        logger.info(f"      ⚠️ 只找到 Barcode: {barcode_only_count} 筆")
        logger.info(f"      ❌ 完全找不到: {unmatched_count} 筆")
        
        return matched_df, unmatched_df


# --- 檔案輸出器 ---
class ReceiptExporter:
    def __init__(self, base_dir: str = "workspace"):
        self.output_root = Path(base_dir) / "output"

        self.pos_dir = self.output_root / "pos_import" # 存放 POS 格式的 XLS
        
        self.pos_dir.mkdir(parents=True, exist_ok=True)

    def save_pos_excel(self, df: pd.DataFrame, original_filename: str):
        """產生 POS 系統專用的舊版 .xls 格式 (字串模式)"""
        # 定義模板 (依照你的需求)
        # 鍵值 (0-12) 對應 Excel 的第 A-M 欄
        Instock_template = {
            0: [
                'MBA POS 入貨表', '', 
                '請不要修改此入貨表之格式!! 如此入貨表之格式被修改, 可能會導致系統不能匯入此表中的資料!!', 
                '系統只會檢查並把 貨號/ 條碼,  入貨價,  入貨量, 店號, 入貨日期, 收據單號 及 供應商編號  資料匯入系統.  貨品及供應商名稱只供客人作參考.', 
                '可在下表 只輸入貨品編號或貨 品條碼.   如在下表同時輸入貨品編號及貨品條碼, 系統會以貨品編號為準.', 
                '請留意, 入貨日期格式為 (年年年年月月日日), 即今天的日期為 20250315', 
                '', '', '', '', '', '', '', '', '貨品編號'
            ],
            1: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '貨品條碼'],
            2: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '貨品名稱'],
            3: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '入貨價'],
            4: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '入貨量'],
            5: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '店號'],
            6: ['Ver:3.2', '', '', '', '', '', '', '', '', '', '', '', '', '', '入貨日期'],
            7: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '收據單號'],
            8: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '供應商編號'],
            9: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '供應商名稱'],
            10: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '備註'],
            11: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '狀態'],
            12: ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '系統備註']
        }

        # 1. 準備 POS 系統要求的欄位順序 (務必與 Template Key 0~12 對應)
        target_columns = [
            '貨品編號', '貨品條碼', '貨品名稱', '入貨價', '入貨量',
            '店號', '入貨日期', '收據單號', '供應商編號', '供應商名稱',
            '備註', '狀態', '系統備註'
        ]

        # 確保 DataFrame 按照這個順序排列，缺少的欄位會在 Cleaner 階段補齊
        # 若有萬一缺少的，這裡補上空字串以免報錯
        for col in target_columns:
            if col not in df.columns:
                df[col] = ''
        
        df_export = df[target_columns].copy()

        # 2. 建立 Excel
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Sheet1')

        # 3. 寫入模板 (Template)
        # Template 結構：Key 是 Column Index，Value 是該 Column 的 Rows List
        for col_idx, row_data_list in Instock_template.items():
            for row_idx, cell_value in enumerate(row_data_list):
                # 強制轉字串
                sheet.write(row_idx, col_idx, str(cell_value))

        # 4. 寫入數據 (Data)
        # 資料從模板最長的那一行之後開始寫入
        start_row = max(len(row) for row in Instock_template.values())
        
        for i, row_data in enumerate(df_export.values):
            for j, cell_value in enumerate(row_data):
                # 強制轉字串 (str)，確保 POS 系統相容性
                sheet.write(start_row + i, j, str(cell_value))

        # 5. 存檔
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"POS_{Path(original_filename).stem}_{timestamp}.xls"
        save_path = self.pos_dir / filename
        
        workbook.save(str(save_path))
        logger.info(f"   💾 POS 匯入檔: {filename}")
    
    def save_unmatched_excel(self, df: pd.DataFrame, supplier_name: str, validator: Optional['ProductValidator'] = None, base_dir: str = "workspace"):
        """
        將找不到對應的產品存成待處理 Excel 檔 (.xlsx)
        
        Args:
            df: 包含「處理原因」欄位的待處理 DataFrame
            supplier_name: 識別到的供應商名稱
            validator: ProductValidator 實例，用於獲取共用條碼的選項
            base_dir: 工作目錄
        """
        if df.empty:
            return
        
        # 確保 pending 資料夾存在
        pending_dir = Path(base_dir) / "pending"
        pending_dir.mkdir(parents=True, exist_ok=True)
        
        # 簡化輸出欄位
        output_columns = [
            '貨品條碼', '貨品名稱', '入貨價', '入貨量', '處理原因', '人手輸入貨品編號'
        ]
        
        # 準備輸出 DataFrame
        # 確保基本欄位存在
        base_cols = ['貨品條碼', '貨品名稱', '入貨價', '入貨量']
        missing_base_cols = [col for col in base_cols if col not in df.columns]
        if missing_base_cols:
            logger.error(f"❌ 待處理檔缺少基本欄位: {missing_base_cols}")
            logger.error(f"   現有欄位: {list(df.columns)}")
            logger.error(f"   DataFrame 行數: {len(df)}")
            raise ValueError(f"缺少基本欄位: {missing_base_cols}")
        
        # 檢查 DataFrame 是否為空（在檢查欄位後）
        if len(df) == 0:
            logger.warning("⚠️ DataFrame 為空，無法保存")
            return
        
        df_export = df[base_cols].copy()
        logger.debug(f"   準備保存 {len(df_export)} 筆記錄")
        
        # 保留或新增「處理原因」欄位
        if '處理原因' in df.columns:
            df_export['處理原因'] = df['處理原因']
            logger.debug(f"   已保留「處理原因」欄位")
        else:
            df_export['處理原因'] = ''
            logger.debug(f"   新增「處理原因」欄位（空值）")
        
        # 「人手輸入貨品編號」欄位：如果是重新保存未填寫的記錄，確保是空白
        # 檢查是否有已填寫的值，如果有則清空（因為這些是未填寫的記錄）
        if '人手輸入貨品編號' in df.columns:
            # 只保留真正為空的記錄（未填寫的）
            df_export['人手輸入貨品編號'] = ''
            logger.debug(f"   已保留「人手輸入貨品編號」欄位（設為空）")
        else:
            df_export['人手輸入貨品編號'] = ''  # 空白欄位，供人工填寫
            logger.debug(f"   新增「人手輸入貨品編號」欄位（空值）")
        
        # 存檔 - 使用新格式：{供應商名稱}需要人手處理{日期}.xlsx
        date_str = datetime.now().strftime("%Y%m%d")
        # 清理供應商名稱，移除可能導致檔案名問題的字元
        safe_supplier_name = supplier_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
        filename = f"{safe_supplier_name}需要人手處理{date_str}.xlsx"
        save_path = pending_dir / filename
        
        try:
            # 先使用 pandas 輸出基本數據
            logger.debug(f"   正在保存到: {save_path}")
            df_export.to_excel(save_path, index=False, engine='openpyxl')
            logger.debug(f"   基本數據已保存，共 {len(df_export)} 筆記錄")
            
            # 使用 openpyxl 添加註解
            try:
                from openpyxl import load_workbook
                from openpyxl.comments import Comment
                
                wb = load_workbook(save_path)
                ws = wb.active
                
                # 找到「人手輸入貨品編號」欄位的索引
                product_code_col_idx = output_columns.index('人手輸入貨品編號') + 1  # Excel 從 1 開始
                
                # 為每一行添加註解（如果是共用條碼）
                # 使用 enumerate 來獲取實際的行號（從 0 開始，對應 Excel 的第 2 行開始，因為第 1 行是標題）
                comment_count = 0
                for excel_row_idx, (df_idx, row) in enumerate(df_export.iterrows(), start=2):
                    barcode = str(row.get('貨品條碼', '')).strip()
                    reason = str(row.get('處理原因', '')).strip()
                    
                    # 如果是共用條碼，添加選項註解
                    if reason == '共用條碼，需人手選擇顏色或大小' and validator:
                        options = validator.get_barcode_options(barcode)
                        if options:
                            # 建立註解內容
                            comment_text = "可選的貨品編號：\n"
                            for opt in options:
                                product_code = opt.get('ProductCode', '')
                                name = opt.get('Name', '')
                                comment_text += f"- {product_code}: {name}\n"
                            
                            # 添加註解到對應的儲存格（excel_row_idx 已經是正確的行號，從 2 開始）
                            cell = ws.cell(row=excel_row_idx, column=product_code_col_idx)
                            cell.comment = Comment(comment_text, "系統")
                            cell.comment.width = 300
                            cell.comment.height = 100
                            comment_count += 1
                
                wb.save(save_path)
                if comment_count > 0:
                    logger.debug(f"   已添加 {comment_count} 個註解")
                logger.info(f"   📋 待處理檔: {filename} (已加入註解)")
            except ImportError:
                logger.warning("   ⚠️ 無法添加註解（需要 openpyxl），但檔案已儲存")
            except Exception as e:
                logger.warning(f"   ⚠️ 添加註解時發生錯誤: {e}，但檔案已儲存")
            
            logger.info(f"   📋 待處理檔: {filename}")
            if '處理原因' in df_export.columns:
                reason_counts = df_export['處理原因'].value_counts().to_dict()
                logger.info(f"      原因統計: {reason_counts}")
            logger.info(f"      檔案路徑: {save_path}")
        except ImportError:
            # 如果沒有 openpyxl，嘗試使用 xlsxwriter（但無法添加註解）
            try:
                df_export.to_excel(save_path, index=False, engine='xlsxwriter')
                logger.warning("   ⚠️ 使用 xlsxwriter 儲存（無法添加註解），建議安裝 openpyxl")
                logger.info(f"   📋 待處理檔: {filename}")
                logger.info(f"      原因統計: {df_export['處理原因'].value_counts().to_dict()}")
            except ImportError:
                logger.error("❌ 需要安裝 openpyxl 或 xlsxwriter 才能輸出 .xlsx 格式")
                logger.info("   請執行: pip install openpyxl")
        except Exception as e:
            logger.error(f"❌ 儲存待處理檔失敗: {e}")
            logger.error(f"   檔案路徑: {save_path}")
            logger.error(f"   資料筆數: {len(df_export)}")
            raise  # 重新拋出異常，讓主流程的 try-except 能捕獲
    
    def process_manual_excel(self, file_path: Path, mapping_manager: 'MappingManager', validator: 'ProductValidator', base_dir: str = "workspace") -> Tuple[pd.DataFrame, int, pd.DataFrame]:
        """
        處理人工填寫的待處理 Excel 檔案
        
        Args:
            file_path: 待處理 Excel 檔案路徑
            mapping_manager: MappingManager 實例
            validator: ProductValidator 實例
            base_dir: 工作目錄
        
        Returns:
            Tuple[pd.DataFrame, int, pd.DataFrame]: 
            - 處理後的 DataFrame（已填寫的記錄）
            - 新增的 mapping 數量
            - 未填寫的 DataFrame（需要保留的記錄）
        """
        try:
            # 讀取 Excel
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
            
            # 檢查必要欄位
            required_cols = ['貨品條碼', '貨品名稱', '人手輸入貨品編號']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                logger.error(f"❌ 待處理檔缺少必要欄位: {missing}")
                return pd.DataFrame(), 0, pd.DataFrame()
            
            # 分離已填寫和未填寫的記錄
            filled_mask = (
                (df['人手輸入貨品編號'].astype(str).str.strip() != '') &
                (df['人手輸入貨品編號'].astype(str).str.strip().str.lower() != 'nan')
            )
            df_filled = df[filled_mask].copy()
            df_unfilled = df[~filled_mask].copy()
            
            if df_filled.empty:
                logger.warning(f"   ⚠️ {file_path.name}: 沒有已填寫的記錄")
                return pd.DataFrame(), 0, df_unfilled
            
            # 從檔名提取供應商名稱
            supplier_name = file_path.stem.split('需要人手處理')[0] if '需要人手處理' in file_path.stem else ''
            
            # 將填寫的記錄加入 mapping
            mapping_count = 0
            processed_rows = []
            
            for idx, row in df_filled.iterrows():
                barcode = str(row['貨品條碼']).strip()
                product_name = str(row['貨品名稱']).strip()
                product_code = str(row['人手輸入貨品編號']).strip()
                
                if barcode and product_name and product_code:
                    mapping_manager.add_mapping(barcode, product_name, product_code, supplier_name)
                    mapping_count += 1
                    
                    # 準備處理後的資料
                    processed_row = {
                        '貨品條碼': barcode,
                        '貨品名稱': product_name,
                        '入貨價': str(row.get('入貨價', '0')).strip(),
                        '入貨量': str(row.get('入貨量', '0')).strip(),
                        '貨品編號': product_code,
                        '供應商名稱': supplier_name,
                        '店號': 'S1',
                        '入貨日期': datetime.now().strftime('%Y%m%d'),
                        '收據單號': '',
                        '供應商編號': '001',
                        '備註': '',
                        '狀態': ''
                    }
                    processed_rows.append(processed_row)
            
            if processed_rows:
                processed_df = pd.DataFrame(processed_rows)
                # 轉換數值欄位
                processed_df['入貨價'] = pd.to_numeric(processed_df['入貨價'], errors='coerce').fillna(0)
                processed_df['入貨量'] = pd.to_numeric(processed_df['入貨量'], errors='coerce').fillna(0).astype(int)
                processed_df = processed_df[processed_df['入貨量'] > 0]
                
                logger.info(f"   ✅ 已處理 {len(processed_df)} 筆產品，新增 {mapping_count} 筆 mapping")
                if not df_unfilled.empty:
                    logger.info(f"   ⚠️ 還有 {len(df_unfilled)} 筆未填寫的記錄需要保留")
                return processed_df, mapping_count, df_unfilled
            else:
                return pd.DataFrame(), 0, df_unfilled
                
        except Exception as e:
            logger.error(f"❌ 處理待處理檔失敗: {e}")
            return pd.DataFrame(), 0, pd.DataFrame()

def main():

    base_dir = "workspace"
    # 1. 讀取設定
    config_mgr = ConfigManager(Path(base_dir))
    config_df = config_mgr.load_config()
    
    # 建立 Mapping 管理器
    mapping_mgr = MappingManager(Path(base_dir))
    
    loader = BatchReceiptLoader(base_dir)
    cleaner = ReceiptCleaner(config_df)
    exporter = ReceiptExporter(base_dir)

    # 讀取並驗證庫存數據源
    input_stock = "data/processed/DetailGoodsStockToday.csv"
    if not os.path.exists(input_stock):
        logger.error("❌ 錯誤: 找不到數據源。")
        return
    
    # 建立產品驗證器（傳入 mapping_manager）
    validator = ProductValidator(input_stock, mapping_mgr)

    logger.info("🚀 開始批次處理...")
    
    # 搜尋關鍵字：從設定檔裡的 4 個關鍵欄位抓字
    search_keywords = []
    if not config_df.empty:
        target_cols = ["貨品條碼", "入貨價", "入貨量", "貨品名稱"]
        for col in target_cols:
            if col in config_df.columns:
                keywords = config_df[col].dropna().unique().tolist()
                search_keywords.extend([k for k in keywords if k.lower() != 'nan' and k.strip()])
    
    # 若 Config 沒東西，給個基本預設值以免程式跑不動
    if not search_keywords:
        logger.warning("⚠️ Config 中無關鍵字，請設定供應商設定檔！")
        return

    # 2. 先處理待處理檔案（人工填寫的）
    logger.info("📋 檢查待處理檔案...")
    manual_files = [f for f in loader.get_pending_files() if '需要人手處理' in f.stem]
    for file_path in manual_files:
        logger.info(f"📝 處理待處理檔: {file_path.name}")
        processed_df, mapping_count, unfilled_df = exporter.process_manual_excel(file_path, mapping_mgr, validator, base_dir)
        
        if not processed_df.empty:
            # 有已填寫記錄 → 需要處理
            # 匯出到 POS 檔
            exporter.save_pos_excel(processed_df, file_path.name)
            logger.info(f"   ✅ 已匯出 {len(processed_df)} 筆產品到 POS 匯入檔")
            
            # 處理未填寫的記錄
            if not unfilled_df.empty:
                # 有未填寫的記錄，重新保存
                supplier_name = file_path.stem.split('需要人手處理')[0] if '需要人手處理' in file_path.stem else ''
                try:
                    exporter.save_unmatched_excel(unfilled_df, supplier_name, validator, base_dir)
                    logger.info(f"   📋 已更新待處理檔，保留 {len(unfilled_df)} 筆未填寫的記錄")
                except Exception as e:
                    logger.error(f"   ❌ 保存未填寫記錄失敗: {e}")
                    logger.warning(f"   ⚠️ 保留原始待處理檔，未歸檔")
                    continue  # 保存失敗時不歸檔，避免遺失資料
            
            # 無論是否有未填寫記錄，都歸檔原始檔（因為已經處理過了）
            loader.archive_file(file_path)
            logger.info(f"   📦 原始待處理檔已歸檔")
        else:
            # 沒有已填寫記錄 → 不用動，保留原檔案
            logger.info(f"   ℹ️ {file_path.name}: 沒有已填寫的記錄，保留原檔案等待處理")
            # 不歸檔，保留在 pending 中
    
    # 3. 處理收據檔案
    logger.info("📄 處理收據檔案...")
    for file_path in loader.get_pending_files():
        # 跳過待處理檔案（已經處理過了）
        if '需要人手處理' in file_path.stem:
            continue
            
        raw_header_df, raw_data_df = loader.smart_load(file_path, search_keywords)
        
        if not raw_data_df.empty:
            clean_df, supplier_name = cleaner.process(raw_data_df)
            
            if clean_df is not None:
                # 產品驗證：分離有對應和找不到的產品（會自動檢查 mapping）
                matched_df, unmatched_df = validator.validate_products(clean_df, supplier_name)
                
                # 處理有對應的產品（正常匯出 POS 檔）
                if not matched_df.empty:
                    exporter.save_pos_excel(matched_df, file_path.name)
                    logger.info(f"   ✅ 已匯出 {len(matched_df)} 筆產品到 POS 匯入檔")
                
                # 處理找不到對應的產品（存待處理檔）
                if not unmatched_df.empty:
                    exporter.save_unmatched_excel(unmatched_df, supplier_name, validator, base_dir)
                    logger.info(f"   ⚠️ 已標記 {len(unmatched_df)} 筆產品待人工處理")
                
                # 所有處理過的原始收據都歸檔到 processed
                loader.archive_file(file_path)
                logger.info(f"   📦 原始收據已歸檔")
            else:
                logger.error("   ❌ 清洗失敗 (未識別或格式錯誤)")
        
    logger.info("程式執行完成，等待用戶確認...")
    logger.info("-" * 30)


# --- 主程式 ---
if __name__ == "__main__":
    main()



